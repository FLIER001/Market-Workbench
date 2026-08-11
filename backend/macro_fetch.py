"""宏观面扩展数据采集 —— 爬取无法通过现成 API 获取的官方统计指标。

数据源全部为公开发布的官方网站，按指标归并为 6 个采集策略：

  1. NBS 国家统计局（data.stats.gov.cn）：
     PMI 分项 / 核心 CPI / 工业企业利润与库存 / 房地产销售与资金来源 /
     固定资产投资分项（设备工器具）/ 财政支出 / 货币供应。
  2. 人民银行（pbc.gov.cn）：
     社会融资规模存量结构（政府债→私人信用脉冲）、金融机构信贷收支（分部门中长期贷款）、
     银行家问卷调查（贷款需求/审批指数，PDF）。
  3. 财政部（mof.gov.cn）：财政收支（一般公共预算 + 政府性基金）。
  4. CPB 荷兰经济政策分析局：世界贸易量。
  5. akshare 衍生：专项债发行（跨市场去重）、中债信用利差、中国货币网资金利率与存单曲线。
  6. 近似代理：非银/财政存款（央行信贷收支表近似）。

所有采集函数返回统一结构：{"label","value","forecast","prev","date","hist","unit","source"}，
与 macro.py 现有 investing 格式指标对齐，便于前端统一渲染。

容错策略：每个函数独立 try/except，失败返回 None，由上层缓存（market.py _sub_cached /
_layered_get）做 last-good 回退，页面不空窗。
"""

from __future__ import annotations

import base64
import io
import json
import math
import re
import time
import zlib
from datetime import datetime, timezone, timedelta

import requests

BEIJING = timezone(timedelta(hours=8))
UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

# ---------------------------------------------------------------------------
# 通用工具
# ---------------------------------------------------------------------------

def _get(url: str, timeout: int = 15, retries: int = 2, **kw) -> requests.Response:
    last = None
    for i in range(retries + 1):
        try:
            r = requests.get(url, headers={"User-Agent": UA}, timeout=timeout, **kw)
            r.raise_for_status()
            return r
        except Exception as e:  # noqa: BLE001
            last = e
            time.sleep(0.8 * (i + 1))
    raise last


def _ym_to_date(ym: str) -> str:
    """'2026年6月' / '2026.06' / '202606' / '2026年第1-2季度' → 'YYYY-MM'。"""
    s = str(ym)
    m = re.match(r"(\d{4})年(\d{1,2})月", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r"(\d{4})\.(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}"
    m = re.match(r"(\d{4})(\d{2})$", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    m = re.match(r"(\d{4})年第?(\d)", s)  # 季度 → 季末月
    if m:
        return f"{m.group(1)}-{int(m.group(2))*3:02d}"
    return s[:10]


def _series_from_rows(date_vals: list[tuple[str, float]], limit: int = 60) -> list[dict]:
    """[(date, value), ...] → 升序 hist 点列。"""
    pts = sorted({d: v for d, v in date_vals if v is not None}.items())
    return [{"date": d, "v": round(float(v), 3)} for d, v in pts[-limit:]]


def _mk(label: str, hist: list[dict], unit: str = "", source: str = "",
        forecast: float | None = None, value_transform=None) -> dict | None:
    """由 hist 点列构建统一指标卡。"""
    if not hist:
        return None
    last = hist[-1]["v"]
    if value_transform:
        last = value_transform(last)
        hist = [{"date": h["date"], "v": value_transform(h["v"])} for h in hist]
    prev = hist[-2]["v"] if len(hist) > 1 else None
    return {
        "label": label,
        "value": round(last, 2),
        "forecast": forecast,
        "prev": round(prev, 2) if prev is not None else None,
        "date": hist[-1]["date"],
        "hist": hist,
        "unit": unit,
        "source": source,
    }


# ---------------------------------------------------------------------------
# 策略 1：NBS 国家统计局
# ---------------------------------------------------------------------------

def _nbs_fetch(kind: str, path: str, period: str = "LAST36"):
    """调 akshare 的 NBS 通用接口，返回 DataFrame（行=指标，列=月份）。"""
    import akshare as ak  # 延迟导入，避免冷启动开销
    try:
        df = ak.macro_china_nbs_nation(kind=kind, path=path, period=period)
    except Exception:
        return None
    if df is None or df.empty:
        return None
    return df


def _nbs_row_hist(df, row_keyword: str, value_transform=None) -> list[dict]:
    """从 NBS DataFrame 里按行名关键词取一条月度序列，转成升序 hist。"""
    if df is None or df.empty:
        return []
    mask = df.index.str.contains(row_keyword, na=False)
    if not mask.any():
        return []
    row = df[mask].iloc[0]
    pts = []
    for col, val in row.items():
        d = _ym_to_date(col)
        try:
            v = float(val)
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        if value_transform:
            v = value_transform(v)
        pts.append((d, v))
    return _series_from_rows(pts)


# ---- 1a. PMI 分项（框架权重最高的缺口，一次性 6 个指标） ----

def pmi_new_orders() -> dict | None:
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    return _mk("PMI 新订单", _nbs_row_hist(df, "新订单指数"), source="统计局 PMI 分项")


def pmi_production() -> dict | None:
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    return _mk("PMI 生产", _nbs_row_hist(df, "生产指数"), source="统计局 PMI 分项")


def pmi_new_export_orders() -> dict | None:
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    return _mk("PMI 新出口订单", _nbs_row_hist(df, "新出口订单指数"), source="统计局 PMI 分项")


def pmi_finished_inventory() -> dict | None:
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    return _mk("PMI 产成品库存", _nbs_row_hist(df, "产成品库存指数"), source="统计局 PMI 分项")


def pmi_output_price() -> dict | None:
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    return _mk("PMI 出厂价格", _nbs_row_hist(df, "出厂价格指数"), source="统计局 PMI 分项")


def pmi_input_price() -> dict | None:
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    return _mk("PMI 购进价格", _nbs_row_hist(df, "主要原材料购进价格指数"), source="统计局 PMI 分项")


def pmi_full() -> dict:
    """一次拉取全部 PMI 分项，返回 {key: 指标卡}，供上层避免重复请求。"""
    df = _nbs_fetch("月度数据", "采购经理指数 > 制造业采购经理指数")
    out = {}
    specs = {
        "pmi_headline": ("制造业PMI", "制造业采购经理指数"),
        "pmi_new_orders": ("PMI 新订单", "新订单指数"),
        "pmi_production": ("PMI 生产", "生产指数"),
        "pmi_new_export_orders": ("PMI 新出口订单", "新出口订单指数"),
        "pmi_finished_inventory": ("PMI 产成品库存", "产成品库存指数"),
        "pmi_output_price": ("PMI 出厂价格", "出厂价格指数"),
        "pmi_input_price": ("PMI 购进价格", "主要原材料购进价格指数"),
        "pmi_employment": ("PMI 从业人员", "从业人员指数"),
        "pmi_expectation": ("PMI 生产经营预期", "生产经营活动预期指数"),
    }
    for key, (label, kw) in specs.items():
        card = _mk(label, _nbs_row_hist(df, kw), source="统计局 PMI 分项")
        if card:
            out[key] = card
    return out


# ---- 1a2. 服务业生产指数 ----

def services_production() -> dict | None:
    """服务业生产指数当月同比（NBS 月度，景气-实际活动模块的服务业验证）。

    NBS 该目录为双层级「服务业生产指数 > 服务业生产指数」。
    """
    df = _nbs_fetch("月度数据", "服务业生产指数 > 服务业生产指数", period="LAST24")
    return _mk("服务业生产指数同比", _nbs_row_hist(df, "服务业生产指数当月同比"),
               unit="%", source="统计局服务业生产指数")


# ---- 1b. 核心 CPI（不包括食品和能源） ----

def core_cpi() -> dict | None:
    """核心 CPI（不包括食品和能源）同比。NBS 分时间段建节点，需拼接。"""
    base = "价格指数 > 居民消费价格分类指数 (上年同月=100) > 全国居民消费价格分类指数 (上年同月=100)"
    pts = []
    for seg in ("(2026-)", "(2021-2025)"):
        df = _nbs_fetch("月度数据", f"{base} {seg}", period="LAST36")
        if df is not None and not df.empty:
            pts.extend(_row_pts(df, "不包括食品和能源", value_transform=lambda v: round(v - 100, 2)))
    hist = _series_from_rows(pts)
    return _mk("核心 CPI 同比", hist, unit="%", source="统计局 CPI 分类")


def _row_pts(df, row_keyword: str, value_transform=None) -> list:
    """从 NBS DataFrame 取一条序列，返回 [(date, value)] 原始点（供拼接）。"""
    if df is None or df.empty:
        return []
    mask = df.index.str.contains(row_keyword, na=False)
    if not mask.any():
        return []
    row = df[mask].iloc[0]
    pts = []
    for col, val in row.items():
        d = _ym_to_date(col)
        try:
            v = float(val)
        except (TypeError, ValueError):
            continue
        if value_transform:
            v = value_transform(v)
        pts.append((d, v))
    return pts


# ---- 1c. 工业企业利润 + 产成品库存（库存周期） ----

def industrial_profit() -> dict | None:
    df = _nbs_fetch("月度数据", "工业 > 工业企业主要经济指标")
    return _mk("工业企业利润累计同比", _nbs_row_hist(df, "利润总额累计增长"),
               unit="%", source="统计局工业企业效益")


def industrial_inventory() -> dict | None:
    df = _nbs_fetch("月度数据", "工业 > 工业企业主要经济指标")
    return _mk("产成品库存同比", _nbs_row_hist(df, "产成品存货增减"),
               unit="%", source="统计局工业企业效益")


def industrial_revenue() -> dict | None:
    df = _nbs_fetch("月度数据", "工业 > 工业企业主要经济指标")
    return _mk("工业企业营收累计同比", _nbs_row_hist(df, "营业收入累计增长"),
               unit="%", source="统计局工业企业效益")


# ---- 1d. 房地产：销售 + 资金来源 ----

def property_sales_area() -> dict | None:
    df = _nbs_fetch("月度数据", "房地产 > 新建商品房销售面积")
    return _mk("新建商品房销售面积累计同比", _nbs_row_hist(df, "销售面积累计增长"),
               unit="%", source="统计局房地产")


def property_funds() -> dict | None:
    df = _nbs_fetch("月度数据", "房地产 > 房地产开发投资实际到位资金")
    return _mk("房地产开发到位资金累计同比", _nbs_row_hist(df, "资金来源小计累计增长"),
               unit="%", source="统计局房地产")


def property_loans() -> dict | None:
    df = _nbs_fetch("月度数据", "房地产 > 房地产开发投资实际到位资金")
    return _mk("房地产国内贷款累计同比", _nbs_row_hist(df, "国内贷款累计增长"),
               unit="%", source="统计局房地产")


# ---- 1e. 固定资产投资分项（设备工器具购置） ----

def fai_equipment() -> dict | None:
    """设备工器具购置投资累计同比。固定资产投资增速节点路径含全角空格，用 cid 直取。"""
    df = _nbs_fetch_by_cid("aac38c7aa152478ebea254ac412aa0a1")
    return _mk("设备工器具购置投资累计同比", _nbs_row_hist(df, "设备工器具购置"),
               unit="%", source="统计局投资")


def _nbs_fetch_by_cid(cid: str, kind: str = "月度数据", period: str = "LAST36"):
    """按目录叶子节点 cid 直接拉 NBS 数据，绕开路径名解析（全角空格等）。"""
    try:
        from akshare.economic.macro_china_nbs import (
            _post_nbs_es_data, _get_nbs_root_id, _get_nbs_indicators,
            _KIND_CONFIG, _build_nbs_dts, _get_nbs_granularity)
    except Exception:
        return None
    import pandas as pd
    try:
        config = _KIND_CONFIG[kind]
        route = str(config["route"])
        gran = _get_nbs_granularity(kind)
        root_id = _get_nbs_root_id(int(config["code"]), route)
        inds = _get_nbs_indicators(cid, route)
        dts = _build_nbs_dts(period=period, granularity=gran)
        data = _post_nbs_es_data(
            cid=cid, root_id=root_id, route=route,
            indicator_ids=[i["_id"] for i in inds],
            das=[{"text": "全国", "value": "000000000000"}],
            show_type="1", dts=dts)
        if not data:
            return None
        # 转成 DataFrame：行=指标 i_showname，列=月份
        recs = {}
        cols = []
        for period_item in data:
            pname = period_item.get("name", "")
            cols.append(pname)
            for v in period_item.get("values", []):
                name = v.get("i_showname") or v.get("_name") or ""
                val = v.get("value")
                recs.setdefault(name, {})[pname] = val
        df = pd.DataFrame(recs).T
        df = df.reindex(columns=cols)
        return df
    except Exception:
        return None


# ---- 1f. 财政支出（NBS 月度） ----

def fiscal_expenditure() -> dict | None:
    df = _nbs_fetch("月度数据", "财政 > 国家财政预算支出")
    return _mk("国家财政支出累计同比", _nbs_row_hist(df, "财政支出.*累计增长|支出.*累计增长"),
               unit="%", source="统计局财政")


# ---------------------------------------------------------------------------
# 策略 2：人民银行
# ---------------------------------------------------------------------------

_PBC_STATS_INDEX = "http://www.pbc.gov.cn/diaochatongjisi/116219/116319/index.html"


def _pbc_latest_year_index() -> str | None:
    """找到最新一年的统计数据汇编页。"""
    try:
        r = _get(_PBC_STATS_INDEX)
        r.encoding = r.apparent_encoding
        m = re.search(r'href="(/diaochatongjisi/116219/116319/(\d{4})ntjsj/index\.html)"', r.text)
        if m:
            return f"http://www.pbc.gov.cn{m.group(1)}"
    except Exception:
        pass
    # 退化为当年
    year = datetime.now(BEIJING).year
    return f"http://www.pbc.gov.cn/diaochatongjisi/116219/116319/{year}ntjsj/index.html"


def _pbc_sub_page(year_index_url: str, sub: str) -> str | None:
    """从年度汇编页找子栏目 URL。"""
    try:
        r = _get(year_index_url)
        r.encoding = r.apparent_encoding
        m = re.search(rf'href="(/diaochatongjisi/116219/116319/\d{{4}}ntjsj/{sub}/index\.html)"', r.text)
        if m:
            return f"http://www.pbc.gov.cn{m.group(1)}"
    except Exception:
        pass
    return None


def _pbc_xlsx_links(page_url: str) -> list[str]:
    """从央行栏目页提取全部 xlsx 附件链接（按 URL 中日期排序，最新在前）。"""
    try:
        r = _get(page_url)
        r.encoding = r.apparent_encoding
        files = re.findall(r'href="(/diaochatongjisi/attachDir/[^"]+\.xlsx?)"', r.text)
        files = sorted(set(files), reverse=True)
        return [f"http://www.pbc.gov.cn{f}" for f in files]
    except Exception:
        return []


def _read_xlsx(url: str):
    import openpyxl
    r = _get(url, timeout=20)
    return openpyxl.load_workbook(io.BytesIO(r.content))


def social_financing_stock() -> dict | None:
    """社融与政府债券存量/增速，用于计算私人信用存量同比。

    央行《社会融资规模存量统计表》xlsx。返回主指标卡（社融存量同比），
    私人部门信用脉冲由上层结合存量结构另行计算。
    """
    year_index = _pbc_latest_year_index()
    if not year_index:
        return None
    sub_url = _pbc_sub_page(year_index, "shrzgm")
    if not sub_url:
        return None
    links = _pbc_xlsx_links(sub_url)
    # 找存量统计表（文件名无序，逐个读表头判断）
    for url in links[:6]:
        try:
            wb = _read_xlsx(url)
        except Exception:
            continue
        ws = wb[wb.sheetnames[0]]
        head = " ".join(str(c.value) for row in ws.iter_rows(min_row=1, max_row=4) for c in row if c.value)
        if "存量" not in head:
            continue
        # 解析：第 5 行是月份表头（2026.1, 2026.2...），第 8 行起是项目
        # 月份在第 5 行，奇数列是存量、偶数列是增速
        months = []
        for c in ws[5]:
            v = c.value
            if v and re.match(r"\d{4}\.\d+", str(v)):
                months.append((c.column, str(v)))
        if not months:
            continue
        # 社融存量/增速行 & 政府债券存量/增速行
        stock_level, stock_growth, gov_level, gov_growth = [], [], [], []
        for row in ws.iter_rows(min_row=6, values_only=False):
            name = str(row[0].value or "")
            if "社会融资规模存量" in name and "AFRE" not in name:
                for col, ym in months:
                    level_cell = ws.cell(row=row[0].row, column=col)
                    gcell = ws.cell(row=row[0].row, column=col + 1)
                    try:
                        stock_level.append((_ym_to_date(ym), float(level_cell.value)))
                    except (TypeError, ValueError):
                        pass
                    try:
                        stock_growth.append((_ym_to_date(ym), float(gcell.value)))
                    except (TypeError, ValueError):
                        pass
            if "政府债券" in name:
                for col, ym in months:
                    level_cell = ws.cell(row=row[0].row, column=col)
                    gcell = ws.cell(row=row[0].row, column=col + 1)
                    try:
                        gov_level.append((_ym_to_date(ym), float(level_cell.value)))
                    except (TypeError, ValueError):
                        pass
                    try:
                        gov_growth.append((_ym_to_date(ym), float(gcell.value)))
                    except (TypeError, ValueError):
                        pass
        hist = _series_from_rows(stock_growth)
        card = _mk("社融存量同比", hist, unit="%", source="人民银行社融存量")
        if card is not None:
            card["stock_level_hist"] = _series_from_rows(stock_level)
            card["gov_bond_growth_hist"] = _series_from_rows(gov_growth)
            card["gov_bond_level_hist"] = _series_from_rows(gov_level)
            card["xlsx_url"] = url
        return card
    return None


def credit_by_sector() -> dict | None:
    """金融机构信贷收支表 → 住户/企事业中长期贷款余额，算同比。

    返回住户中长期贷款（主卡）+ 企事业中长期贷款序列（附加），
    以及非银存款、财政性存款（策略 6 近似用）。
    """
    year_index = _pbc_latest_year_index()
    if not year_index:
        return None
    sub_url = _pbc_sub_page(year_index, "jrjgxdsztj")
    if not sub_url:
        return None
    links = _pbc_xlsx_links(sub_url)
    for url in links:
        try:
            wb = _read_xlsx(url)
        except Exception:
            continue
        ws = wb[wb.sheetnames[0]]
        head = " ".join(str(c.value) for row in ws.iter_rows(min_row=1, max_row=8) for c in row if c.value)
        # 只取全口径「（存款类）金融机构本外币信贷收支表」，跳过中资大型/中小型银行等子口径
        if "本外币信贷收支" not in head or "中资" in head:
            continue
        # 月份表头自动定位：前 10 行中含 '2026.01' 样式单元格的那一行
        months = []
        for hdr_row in range(1, 11):
            row_months = []
            for c in ws[hdr_row]:
                v = c.value
                if v and re.match(r"\d{4}\.\d+", str(v)):
                    row_months.append((c.column, str(v)))
            if len(row_months) >= 3:
                months = row_months
                break
        if not months:
            continue
        # 行定位：按项目名
        rows_map = {}
        for row in ws.iter_rows(min_row=7, values_only=False):
            name = str(row[0].value or "")
            for key, kw in [("household_ml", "中长期贷"), ("corp_total", "企（"), ]:
                pass
            # 精确匹配有难度，宽松收集关键行
            label = name.replace(" ", "").replace("　", "")
            if label:
                rows_map.setdefault(label, row[0].row)
        # 按区块状态机定位：住户/企事业各自的中长期贷款、票据融资、财政性存款、非银存款
        def _cells(row_idx: int) -> list[tuple[str, float]]:
            pts = []
            for col, ym in months:
                try:
                    pts.append((_ym_to_date(ym), float(ws.cell(row=row_idx, column=col).value)))
                except (TypeError, ValueError):
                    pass
            return pts

        household = []          # 住户中长期贷款
        corp = []               # 企事业中长期贷款
        corp_total = []         # 企事业贷款总额
        bill_financing = []     # 票据融资
        fiscal_dep = []         # 财政性存款
        nonbank_dep = []        # 非银行业金融机构存款
        section = None          # 当前所在区块
        for row in ws.iter_rows(min_row=7, values_only=False):
            name = re.sub(r"\s+", "", str(row[0].value or ""))
            rid = row[0].row
            if "住户贷款" in name:
                section = "household"
            elif "企（事）业单位贷款" in name or "企事业单位贷款" in name:
                section = "corp"
                corp_total = _cells(rid)
            elif "非银行业金融机构贷款" in name:
                section = "nonbank_loan"
            elif "境外贷款" in name or "债券投资" in name:
                section = None
            elif ("非银行业金融机构存款" in name or "非银行存款" in name
                    or "非存款类金融机构存款" in name):
                nonbank_dep = _cells(rid)
            elif "财政性存款" in name:
                fiscal_dep = _cells(rid)
            if "中长期贷款" in name:
                if section == "household":
                    household = _cells(rid)
                elif section == "corp":
                    corp = _cells(rid)
            if "票据融资" in name and section == "corp":
                bill_financing = _cells(rid)
        hist = _series_from_rows(household)
        card = _mk("住户中长期贷款余额", hist, unit="亿元", source="人民银行信贷收支")
        if card is not None:
            card["corp_ml_loan_hist"] = _series_from_rows(corp)      # 企事业中长期贷款
            card["corp_total_loan_hist"] = _series_from_rows(corp_total)
            card["bill_financing_hist"] = _series_from_rows(bill_financing)  # 票据融资
            card["fiscal_deposit_hist"] = _series_from_rows(fiscal_dep)
            card["nonbank_deposit_hist"] = _series_from_rows(nonbank_dep)
            card["xlsx_url"] = url
        return card
    return None


def bank_survey() -> dict | None:
    """银行家问卷调查：贷款总体需求指数 + 贷款审批指数（季度，PDF 解析）。"""
    try:
        from pypdf import PdfReader
    except ImportError:
        return None
    # 问卷列表页
    list_url = "http://www.pbc.gov.cn/diaochatongjisi/116219/116227/index.html"
    try:
        r = _get(list_url)
        r.encoding = r.apparent_encoding
        # 找最新银行家问卷调查报告页
        pages = re.findall(r'href="(/diaochatongjisi/116219/116227/[^"]+/index\.html)"[^>]*>([^<]*银行家问卷调查报告[^<]*)<', r.text)
        if not pages:
            return None
        page_url = f"http://www.pbc.gov.cn{pages[0][0]}"
        r2 = _get(page_url)
        r2.encoding = r2.apparent_encoding
        pdfs = re.findall(r'href="([^"]+\.pdf)"', r2.text)
        if not pdfs:
            return None
        pdf_url = pdfs[0]
        if not pdf_url.startswith("http"):
            pdf_url = f"http://www.pbc.gov.cn{pdf_url}"
        rp = _get(pdf_url, timeout=25)
        reader = PdfReader(io.BytesIO(rp.content))
        text = ""
        for page in reader.pages[:5]:
            text += (page.extract_text() or "") + "\n"
        # 解析：贷款总体需求指数为 X%；审批指数
        demand = re.search(r"贷款总体需求指数为?\s*([\d.]+)\s*%", text)
        approve = re.search(r"(?:贷款)?审批指数为?\s*([\d.]+)\s*%", text)
        quarter = re.search(r"(\d{4})年(?:第)?([一二三四1-4])季度", pages[0][1] or text)
        val = float(demand.group(1)) if demand else None
        appr = float(approve.group(1)) if approve else None
        if val is None:
            return None
        date_str = ""
        if quarter:
            qmap = {"一": 3, "二": 6, "三": 9, "四": 12, "1": 3, "2": 6, "3": 9, "4": 12}
            date_str = f"{quarter.group(1)}-{qmap.get(quarter.group(2), 12):02d}"
        return {
            "label": "银行贷款需求指数",
            "value": val,
            "forecast": None,
            "prev": None,
            "date": date_str or datetime.now(BEIJING).strftime("%Y-%m"),
            "hist": [{"date": date_str, "v": val}] if date_str else [],
            "unit": "%",
            "source": "人民银行银行家问卷调查",
            "approve_index": appr,
            "pdf_url": pdf_url,
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 策略 3：财政部财政收支
# ---------------------------------------------------------------------------

def fiscal_revenue_expenditure() -> dict | None:
    """财政部月度财政收支：两本账支出合成的广义财政支出同比。"""
    index_url = "http://www.mof.gov.cn/zhengwuxinxi/caizhengshuju/"
    try:
        r = _get(index_url)
        r.encoding = r.apparent_encoding
        # 最新「财政收支情况」
        links = re.findall(r'href="(http://gks\.mof\.gov\.cn/tongjishuju/[^"]+\.htm)"[^>]*>([^<]*财政收支情况[^<]*)<', r.text)
        if not links:
            # 链接文字可能在下一行
            links = re.findall(r'href="(http://gks\.mof\.gov\.cn/tongjishuju/[^"]+\.htm)"[^>]*>\s*([^<]*财政收支[^<]*)', r.text)
        if not links:
            return None
        page_url = links[0][0]
        r2 = _get(page_url)
        r2.encoding = r2.apparent_encoding
        body = re.sub(r"<[^>]+>", " ", r2.text)
        body = re.sub(r"\s+", " ", body)
        def _budget(pattern: str):
            m = re.search(pattern + r"\s*([\d.]+)\s*亿元[^。]{0,30}?同比(?:增长|下降)\s*([\d.]+)\s*%", body)
            if not m:
                return None
            sign_text = m.group(0)
            yoy = float(m.group(2)) * (-1 if "下降" in sign_text else 1)
            return float(m.group(1)), yoy

        general = _budget(r"(?:全国)?一般公共预算支出")
        fund = _budget(r"(?:全国)?政府性基金预算支出")
        # 期间（上半年/1-5月等）
        period_m = re.search(r"(上半年|一季度|1-(\d+)月|(\d+)月份)", body)
        if not general:
            return None
        current_total = general[0] + (fund[0] if fund else 0.0)
        prior_total = general[0] / (1.0 + general[1] / 100.0)
        if fund and 1.0 + fund[1] / 100.0 > 0:
            prior_total += fund[0] / (1.0 + fund[1] / 100.0)
        broad_yoy = (current_total / prior_total - 1.0) * 100.0 if prior_total else None
        month_str = ""
        if period_m:
            p = period_m.group(0)
            if "上半年" in p or "1-6月" in p:
                month_str = f"{datetime.now(BEIJING).year}-06"
            elif "一季度" in p or "1-3月" in p:
                month_str = f"{datetime.now(BEIJING).year}-03"
            else:
                m2 = re.search(r"1-(\d+)月", p)
                if m2:
                    month_str = f"{datetime.now(BEIJING).year}-{int(m2.group(1)):02d}"
        return {
            "label": "广义财政支出同比(两本账)",
            "value": round(broad_yoy, 2) if broad_yoy is not None else general[1],
            "forecast": None,
            "prev": None,
            "date": month_str or datetime.now(BEIJING).strftime("%Y-%m"),
            "hist": [],
            "unit": "%",
            "source": "财政部·一般公共预算+政府性基金预算支出",
            "general_exp": {"amount": general[0], "yoy": general[1]},
            "fund_exp": {"amount": fund[0], "yoy": fund[1]} if fund else None,
            "page_url": page_url,
        }
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 策略 4：CPB 世界贸易量
# ---------------------------------------------------------------------------

def world_trade_volume() -> dict | None:
    """CPB World Trade Monitor：世界贸易量指数（2021=100，季调）。"""
    try:
        import pandas as pd
    except ImportError:
        return None
    try:
        r = _get("https://www.cpb.nl/en/worldtrademonitor/latest")
        m = re.search(r'href="(/system/files/cpbmedia/CPB-World-trade-monitor-[^"]+\.xlsx)"', r.text)
        if not m:
            return None
        xlsx_url = f"https://www.cpb.nl{m.group(1)}"
        rx = _get(xlsx_url, timeout=30)
        df = pd.read_excel(io.BytesIO(rx.content), sheet_name="trade_out", header=None)
        # 月份表头行：含 '2000m01' 的那一行
        hdr = None
        for i in range(min(12, len(df))):
            row = df.iloc[i].astype(str)
            if row.str.contains(r"\d{4}m\d{2}", regex=True).any():
                hdr = df.iloc[i]
                break
        if hdr is None:
            return None
        # World trade 数据行
        wt = None
        for i in range(len(df)):
            if str(df.iloc[i, 1]).strip() == "World trade":
                wt = df.iloc[i]
                break
        if wt is None:
            return None
        pts = []
        for j in range(len(hdr)):
            cell = str(hdr[j])
            ym = re.match(r"(\d{4})m(\d{2})", cell)
            if not ym:
                continue
            try:
                v = float(wt[j])
            except (TypeError, ValueError):
                continue
            pts.append((f"{ym.group(1)}-{ym.group(2)}", v))
        hist = _series_from_rows(pts)
        return _mk("世界贸易量指数", hist, source="CPB World Trade Monitor")
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 策略 5：akshare 衍生（专项债 / 已有宏观补充）
# ---------------------------------------------------------------------------

def special_bond_issuance() -> dict | None:
    """新增专项债近3月月均发行额（亿元），跨市场去重且排除未完结月份。

    注意：bond_local_government_issue_cninfo 默认 start/end_date 写死在 2021-09~2021-11，
    必须显式传日期区间，否则永远只返回 2021 年存量数据。这里取近 24 个月滚动窗口，
    逐年分段拉取（接口单段跨度有限），再按月汇总。
    """
    import akshare as ak
    frames = []
    now = datetime.now(BEIJING)
    # 近 3 个自然年，确保覆盖近 24 个月
    for yr in range(now.year - 2, now.year + 1):
        sd, ed = f"{yr}0101", f"{yr}1231"
        try:
            f = ak.bond_local_government_issue_cninfo(start_date=sd, end_date=ed)
            if f is not None and not f.empty:
                frames.append(f)
        except Exception:
            continue
    if not frames:
        return None
    import pandas as pd
    df = pd.concat(frames, ignore_index=True)
    if df is None or df.empty:
        return None
    try:
        name_col = "债券名称" if "债券名称" in df.columns else "债券简称"
        date_col = "发行起始日"
        amt_col = "实际发行总量"
        df = df.copy()
        df[date_col] = pd_to_datetime(df[date_col])
        # 同一只债通常以银行间/上交所/深交所三个代码重复出现，按名称+日期+金额去重。
        df = df.drop_duplicates(subset=[name_col, date_col, amt_col])
        mask = (df[name_col].str.contains("专项", na=False)
                & ~df[name_col].str.contains("再融资|置换", na=False))
        sub = df[mask].dropna(subset=[date_col])
        monthly = sub.groupby(sub[date_col].dt.to_period("M"))[amt_col].sum().sort_index()
        if monthly.empty:
            return None
        monthly = monthly.asfreq("M", fill_value=0.0)
        current_month = pd.Period(datetime.now(BEIJING).strftime("%Y-%m"), freq="M")
        monthly = monthly[monthly.index < current_month]
        rolling = monthly.rolling(3, min_periods=3).mean().dropna()
        pts = [(str(p), float(v)) for p, v in rolling.items()]
        hist = _series_from_rows(pts, limit=36)
        card = _mk("新增专项债近3月月均发行额", hist, unit="亿元", source="巨潮资讯·地方政府债发行")
        if card:
            card["source_url"] = "https://www.cninfo.com.cn/new/commonUrl?url=disclosure/bond/notice"
        return card
    except Exception:
        return None


def pd_to_datetime(s):
    import pandas as pd
    return pd.to_datetime(s, errors="coerce")


# ---------------------------------------------------------------------------
# 金融条件：DR007-政策利率 / AAA 同业存单利差 / AAA 信用利差
# ---------------------------------------------------------------------------

def _reverse_repo_rate_schedule() -> list[tuple[str, float]]:
    """读取人民银行利率调整公告，返回 [(生效日, 7天逆回购利率), ...]。"""
    index_url = "https://www.pbc.gov.cn/zhengcehuobisi/125207/125213/125431/125469/index.html"
    try:
        r = _get(index_url)
        r.encoding = r.apparent_encoding
        links = re.findall(r'href="([^"]+/index\.html)"[^>]*>[^<]*公开市场业务公告[^<]*<', r.text)
    except Exception:
        return []
    out: list[tuple[str, float]] = []
    for path in list(dict.fromkeys(links))[:20]:
        try:
            page_url = f"https://www.pbc.gov.cn{path}" if path.startswith("/") else path
            rp = _get(page_url)
            rp.encoding = rp.apparent_encoding
            zoom = re.search(r'<div[^>]+id="zoom"[^>]*>(.*?)</div>', rp.text, re.S | re.I)
            body = re.sub(r"\s+", "", re.sub(r"<[^>]+>", " ", zoom.group(1) if zoom else rp.text))
            m = re.search(
                r"从(?:(\d{4})年)?(\d{1,2})月(\d{1,2})日起.*?"
                r"7天期逆回购操作利率由此前的?([\d.]+)%调整为([\d.]+)%", body)
            if not m:
                continue
            # 老公告常只写“9月27日”；当前窗口始于2026年，只需保留明确写出年份的调整。
            if not m.group(1):
                continue
            year = int(m.group(1))
            date = f"{year:04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            out.append((date, float(m.group(5))))
            # 最早一条公告同时给出调整前利率，补成时间表的左端点。
            out.append((f"{year - 1:04d}-01-01", float(m.group(4))))
        except Exception:
            continue
    return sorted(dict(out).items())


def _policy_rate_on(date: str, schedule: list[tuple[str, float]]) -> float | None:
    rate = None
    for effective, value in schedule:
        if effective <= date:
            rate = value
        else:
            break
    return rate


def financial_conditions() -> dict:
    """三项日度金融条件；每个子源独立降级，返回可直接展开的指标卡字典。"""
    import akshare as ak

    now = datetime.now(BEIJING)
    start = now - timedelta(days=140)
    schedule = _reverse_repo_rate_schedule()
    out: dict = {}

    # DR007：银行间存款类机构质押式回购定盘利率（FDR007）减政策利率。
    try:
        repo = ak.repo_rate_hist(start_date=start.strftime("%Y%m%d"), end_date=now.strftime("%Y%m%d"))
        pts = []
        for _, row in repo.iterrows():
            date = str(row.get("date", ""))[:10]
            policy = _policy_rate_on(date, schedule)
            if policy is not None and row.get("FDR007") is not None:
                pts.append((date, (float(row["FDR007"]) - policy) * 100.0))
        card = _mk("DR007-7天逆回购利差", _series_from_rows(pts, limit=120), unit="bp",
                   source="中国货币网FDR007+人民银行政策利率")
        if card:
            card["source_url"] = "https://www.chinamoney.com.cn/chinese/bkfrr/"
            out["dr007_policy_spread"] = card
    except Exception:
        pass

    # 中债 AAA 中短期票据 3Y - 国债 3Y。
    try:
        curves = ak.bond_china_yield(start_date=start.strftime("%Y%m%d"), end_date=now.strftime("%Y%m%d"))
        gov = curves[curves["曲线名称"] == "中债国债收益率曲线"]
        aaa = curves[curves["曲线名称"] == "中债中短期票据收益率曲线(AAA)"]
        gm = {str(r["日期"]): float(r["3年"]) for _, r in gov.iterrows() if r.get("3年") == r.get("3年")}
        pts = [(str(r["日期"]), (float(r["3年"]) - gm[str(r["日期"])]) * 100.0)
               for _, r in aaa.iterrows()
               if r.get("3年") == r.get("3年") and str(r["日期"]) in gm]
        card = _mk("AAA信用利差(3年)", _series_from_rows(pts, limit=120), unit="bp",
                   source="中债AAA中短票3Y-国债3Y")
        if card:
            card["source_url"] = "https://yield.chinabond.com.cn/"
            out["credit_spread_aaa"] = card
    except Exception:
        pass

    # 中国货币网 AAA 同业存单 1Y - 7天逆回购政策利率。接口单次仅返回50行，分窗取数。
    try:
        ncd: dict[str, float] = {}
        for offset in range(0, 84, 14):
            end = now - timedelta(days=offset)
            begin = end - timedelta(days=13)
            df = ak.bond_china_close_return(
                symbol="同业存单(AAA)", period="1",
                start_date=begin.strftime("%Y%m%d"), end_date=end.strftime("%Y%m%d"))
            for _, row in df.iterrows():
                if row.get("期限") == row.get("期限") and abs(float(row["期限"]) - 1.0) < 1e-6:
                    ncd[str(row["日期"])] = float(row["到期收益率"])
        pts = []
        for date, value in sorted(ncd.items()):
            policy = _policy_rate_on(date, schedule)
            if policy is not None:
                pts.append((date, (value - policy) * 100.0))
        card = _mk("AAA同业存单1Y-政策利率", _series_from_rows(pts, limit=60), unit="bp",
                   source="中国货币网同业存单(AAA)收益率曲线")
        if card:
            card["source_url"] = "https://www.chinamoney.com.cn/chinese/bkcurvclosedyhis/"
            out["ncd_aaa_spread"] = card
    except Exception:
        pass
    return out


# ---------------------------------------------------------------------------
# 近似指标（可近似项，用免费稳定序列替代付费源）
# ---------------------------------------------------------------------------

def copper_oil_ratio() -> dict | None:
    """铜油比（FRED PCOPPUSDM 铜 ÷ POILWTIUSDM WTI 油，月度）。

    全球增长相对通胀代理：上升=增长强于通胀压力，利多风险资产。
    """
    import market
    copper = market._fred_series_cached("PCOPPUSDM", 60)
    oil = market._fred_series_cached("POILWTIUSDM", 60)
    if not copper or not oil:
        return None
    om = {d: v for d, v in oil}
    pts = []
    for d, cv in copper:
        if d in om and om[d]:
            pts.append((d[:7], round(cv / om[d], 4)))
    hist = _series_from_rows(pts)
    return _mk("铜油比", hist, source="FRED 铜/油")


def resale_house_breadth() -> dict | None:
    """二手房价格动量（京沪二手住宅环比均值代理）。

    akshare 该接口仅含上海/北京两城（非 70 城全集），京沪是二手房市场风向标，
    用两城二手住宅环比指数均值（>100=上涨）近似价格扩散方向。
    精确的 70 城扩散度需逐城爬取统计局分城市数据（70 次/月），本期未做。
    """
    import akshare as ak
    try:
        df = ak.macro_china_new_house_price()
    except Exception:
        return None
    if df is None or df.empty:
        return None
    try:
        col = "二手住宅价格指数-环比"
        df = df.dropna(subset=[col])
        pts = []
        for date, grp in df.groupby("日期"):
            vals = grp[col].dropna().tolist()
            if not vals:
                continue
            pts.append((str(date)[:7], round(sum(vals) / len(vals), 2)))
        hist = _series_from_rows(pts)
        return _mk("二手房价格环比(京沪)", hist, source="统计局房价(京沪代理)")
    except Exception:
        return None


# ---------------------------------------------------------------------------
# 原财新 / RatingDog PMI：S&P Global 连续序列 + 多源回退
# ---------------------------------------------------------------------------

_TE_PMI_PAGES = {
    "chinamanpmi": "https://tradingeconomics.com/china/manufacturing-pmi",
    "chinaserpmi": "https://tradingeconomics.com/china/services-pmi",
}


def _te_decode_chart(payload: str, key: str) -> dict | list:
    """解码 Trading Economics 官网图表使用的公开压缩载荷。

    解码方法来自其网页端 ``ec.min.js``：Base64 → 与页面公开 key 逐字节异或
    → zlib inflate。key 和图表访问 token 均从指标页实时读取，不在代码中固化。
    """
    raw = base64.b64decode(payload)
    key_bytes = key.encode("utf-8")
    if not key_bytes:
        raise ValueError("empty Trading Economics chart key")
    decoded = bytes(v ^ key_bytes[i % len(key_bytes)] for i, v in enumerate(raw))
    # CloudFront 当前可能返回 zlib 或 gzip 包装；MAX_WBITS|32 自动识别两者。
    return json.loads(zlib.decompress(decoded, zlib.MAX_WBITS | 32).decode("utf-8"))


def _te_chart_points(doc: dict | list, limit: int = 60) -> list[dict]:
    """从 Trading Economics 图表 JSON 中提取月度 ``date/value`` 序列。"""
    try:
        series = doc[0]["series"][0]["serie"]
        if str(series.get("source") or "").lower() != "s&p global":
            return []
        rows = series.get("data") or []
    except (IndexError, KeyError, TypeError):
        return []
    pts = []
    for row in rows:
        try:
            if isinstance(row, dict):
                value = row.get("y")
                date = row.get("date")
            else:
                value, date = row[0], row[3]
            if value is None or not date:
                continue
            value = float(value)
            if not math.isfinite(value):
                continue
            pts.append((str(date)[:7], value))
        except (IndexError, TypeError, ValueError):
            continue
    return _series_from_rows(pts, limit=limit)


def _tradingeconomics_pmi_hist(symbol: str, limit: int = 60) -> list[dict]:
    """读取 S&P Global 编制的原财新/现 RatingDog PMI 连续序列。

    Trading Economics 指标页公开展示并标注数据源为 S&P Global。先从页面动态
    读取图表域名、访问 token 与解码 key，再请求同页图表数据，避免把网页构建期
    参数写死。该源覆盖 2025-07 品牌切换前后的同口径历史序列。
    """
    page_url = _TE_PMI_PAGES[symbol]
    html = _get(page_url, timeout=20).text

    def _cfg(name: str) -> str:
        m = re.search(rf"\b{name}\s*=\s*['\"]([^'\"]+)", html)
        if not m:
            raise ValueError(f"Trading Economics page missing {name}")
        return m.group(1)

    datasource = _cfg("TEChartsDatasource").rstrip("/")
    token = _cfg("TEChartsToken")
    key = _cfg("TEObfuscationkey")
    span_years = max(5, math.ceil(limit / 12))
    url = f"{datasource}/economics/{symbol}?span={span_years}y"
    r = requests.get(
        url,
        headers={"User-Agent": UA, "x-api-key": token, "Referer": page_url},
        timeout=20,
    )
    r.raise_for_status()
    payload = r.json()
    if not isinstance(payload, str):
        raise ValueError("unexpected Trading Economics chart payload")
    return _te_chart_points(_te_decode_chart(payload, key), limit=limit)

def _jin10_hist(attr_id: str, limit: int = 60):
    """金十数据中心历史序列（公开 API，attr_id: 73=财新制造业, 67=财新服务业）。

    返回 (hist_list, forecast)；今值为空的行（未发布）已剔除。
    """
    headers = {
        "user-agent": UA,
        "x-app-id": "rU6QIu7JHe2gOUeR",
        "x-csrf-token": "x-csrf-token",
        "x-version": "1.0.0",
    }
    params = {"max_date": "", "category": "ec", "attr_id": attr_id,
              "_": str(int(time.time() * 1000))}
    r = requests.get("https://datacenter-api.jin10.com/reports/list_v2",
                     params=params, headers=headers, timeout=20)
    values = (r.json().get("data") or {}).get("values") or []
    rows = [(v[0], v[1], v[2]) for v in values
            if len(v) >= 2 and v[1] not in (None, "")]
    rows.sort(key=lambda x: x[0])
    rows = rows[-limit:]
    # 保留金十发布日与今值，由 _caixin_pmi 对齐到数据月。
    hist = [(str(d)[:10], round(float(val), 2)) for d, val, _ in rows]
    fc = None
    if rows:
        raw_fc = rows[-1][2]
        if raw_fc not in (None, ""):
            try:
                fc = round(float(raw_fc), 2)
            except (TypeError, ValueError):
                fc = None
    return hist, fc


def _caixin_web_latest() -> dict:
    """财新官网首页解析最新一期制造业/服务业 PMI 数据月值。

    首页「财智研究」栏固定摘要最新一期（"YYYY年M月，财新中国制造业PMI录得X，
    财新中国服务业PMI录得Y"），直接给出两个数据月值；缺失时再扫头条标题
    「N月财新中国制造业/服务业PMI录得X」补齐（按数据月归属，避免 1 月数据
    与 12 月数据同月发布时误取）。

    返回 {"YYYY-MM": {"manufacturing": v, "services": v}}。
    """
    html = _get("https://pmi.caixin.com/", timeout=20).text
    out: dict = {}

    # 财智研究栏：一次给两个数据月值
    i = html.find("财智研究")
    if i != -1:
        blk = html[i: html.find("财智研究end") if "财智研究end" in html else i + 2000]
        m = re.search(r"<p><a href=[^>]*>([^<]+)", blk)
        if m:
            sn = m.group(1)
            ym = re.search(r"(\d{4})年(\d{1,2})月", sn)
            man = re.search(r"财新中国制造业PMI录得(\d+\.?\d*)", sn)
            ser = re.search(r"财新中国服务业PMI录得(\d+\.?\d*)", sn)
            if ym and (man or ser):
                key = f"{ym.group(1)}-{int(ym.group(2)):02d}"
                rec = out.setdefault(key, {})
                if man:
                    rec["manufacturing"] = float(man.group(1))
                if ser:
                    rec["services"] = float(ser.group(1))

    # 头条标题兜底：N月财新中国XPMI…X
    pat = re.compile(
        r"(\d{1,2})月财新中国(制造业|服务业)PMI[^0-9]{0,12}?(\d{2}\.\d)")
    links = re.findall(
        r'<a href="https://pmi\.caixin\.com/(\d{4})-(\d{2})-\d{2}/\d+\.html">([^<]+)</a>',
        html)
    for yy, mm, title in links:
        tm = pat.search(title)
        if not tm:
            continue
        month, kind, val = int(tm.group(1)), tm.group(2), float(tm.group(3))
        data_year = int(yy) if month <= int(mm) else int(yy) - 1
        key = f"{data_year}-{month:02d}"
        rec = out.setdefault(key, {})
        field = "manufacturing" if kind == "制造业" else "services"
        rec.setdefault(field, val)
    return out


def _caixin_pmi(attr_id: str, label: str, field: str) -> dict | None:
    """原财新/现 RatingDog PMI 通用卡。

    2025-07 后财新停止冠名，同一套 S&P Global 调查更名为 RatingDog 中国 PMI。
    主源采用 Trading Economics 公开展示、来源标注为 S&P Global 的连续序列；
    财新官网与金十用于旧品牌时期历史回填和主源故障回退。
    """
    try:
        jin_rows, fc = _jin10_hist(attr_id)
    except Exception:
        jin_rows, fc = [], None
    try:
        web = _caixin_web_latest()
    except Exception:
        web = {}
    symbol = "chinamanpmi" if field == "manufacturing" else "chinaserpmi"
    try:
        te_hist = _tradingeconomics_pmi_hist(symbol)
    except Exception:
        te_hist = []

    web_pts = {ym: rec[field] for ym, rec in web.items() if rec.get(field) is not None}

    def _data_month(pub: str) -> str:
        y, m, d = int(pub[:4]), int(pub[5:7]), int(pub[8:10])
        if d >= 28:  # 月末提前发布 → 数据月=当月
            return f"{y}-{m:02d}"
        m -= 1  # 月初发布 → 数据月=上一月
        if m == 0:
            y, m = y - 1, 12
        return f"{y}-{m:02d}"

    merged: dict = {}
    for pub, v in jin_rows:
        merged[_data_month(pub)] = v
    # 官网数据月覆盖/续接（权威）
    for ym, v in web_pts.items():
        merged[ym] = v
    # S&P Global 连续序列覆盖品牌切换前后；当前值以该源为准。
    for point in te_hist:
        merged[point["date"]] = point["v"]

    if not merged:
        return None
    hist = [{"date": k, "v": round(v, 2)} for k, v in sorted(merged.items())][-60:]
    source = ("S&P Global（Trading Economics）+财新/金十历史"
              if te_hist else "财新官网+金十（更新滞后）")
    card = _mk(label, hist, source=source)
    # 金十的预期值只在它与最新实际值属于同一数据月时才可使用，防止把
    # 2025 年旧预期挂到 2026 年 RatingDog 最新值旁边。
    jin_latest_month = _data_month(jin_rows[-1][0]) if jin_rows else None
    if card is not None and fc is not None and jin_latest_month == card["date"]:
        card["forecast"] = fc
    return card


def caixin_manufacturing_pmi() -> dict | None:
    """RatingDog 制造业 PMI（原财新，S&P Global 连续口径）。"""
    return _caixin_pmi("73", "RatingDog制造业 PMI（原财新）", "manufacturing")


def caixin_services_pmi() -> dict | None:
    """RatingDog 服务业 PMI（原财新，S&P Global 连续口径）。"""
    return _caixin_pmi("67", "RatingDog服务业 PMI（原财新）", "services")


def usdcnh() -> dict | None:
    """人民币汇率（美元中间价，currency_boc_safe 中国外汇交易中心，8000+ 行稳定）。

    注：东财 USDCNH 离岸接口（forex_hist_em）经 push2his 在当前网络环境不稳定，
    用央行中间价（在岸指导价）代理人民币汇率压力方向，离岸/在岸价差不可得。
    中间价单位：100 美元兑人民币（678.94 ≈ USDCNY 6.79）。
    """
    import akshare as ak
    try:
        df = ak.currency_boc_safe()
    except Exception:
        return None
    if df is None or df.empty:
        return None
    try:
        pts = []
        for _, r in df.tail(250).iterrows():
            try:
                pts.append((str(r["日期"])[:10], round(float(r["美元"]) / 100, 4)))
            except (TypeError, ValueError):
                continue
        hist = _series_from_rows(pts)
        return _mk("人民币汇率(美元中间价)", hist, source="外汇交易中心中间价")
    except Exception:
        return None


def nonbank_deposit() -> dict | None:
    """非银存款（非存款类金融机构存款余额，信贷收支表）。"""
    card = credit_by_sector()
    if not card:
        return None
    h = card.get("nonbank_deposit_hist") or []
    if not h:
        return None
    return {
        "label": "非银存款余额", "value": h[-1]["v"], "forecast": None,
        "prev": h[-2]["v"] if len(h) > 1 else None,
        "date": h[-1]["date"], "hist": h, "unit": "亿元",
        "source": "人民银行信贷收支",
    }


def policy_execution() -> dict | None:
    """政策执行确认合成：财政支出同比 + 专项债发行 + 企事业中长期贷款三线。

    取三者方向的简单合成（改善数），0-3 分，越高=执行越实。
    """
    import market
    ind = {}
    try:
        # 财政支出同比
        fe = fiscal_revenue_expenditure()
        if fe and fe.get("value") is not None:
            ind["fiscal"] = fe["value"]
        # 专项债发行（近3月均值 vs 前3月）
        sb = special_bond_issuance()
        if sb and len(sb.get("hist", [])) >= 6:
            h = sb["hist"]
            recent = sum(p["v"] for p in h[-3:]) / 3
            prior = sum(p["v"] for p in h[-6:-3]) / 3
            ind["bond_momentum"] = round((recent - prior) / max(prior, 1) * 100, 1)
        # 企事业中长期贷款动量
        cbs = credit_by_sector()
        if cbs and len(cbs.get("corp_ml_loan_hist", [])) >= 4:
            h = cbs["corp_ml_loan_hist"]
            ind["corp_loan_chg"] = round(h[-1]["v"] - h[-4]["v"], 0)
    except Exception:
        pass
    if not ind:
        return None
    # 合成：财政>0、专项债动量>0、企业贷款增 → 各记 1 分
    score = 0
    parts = []
    if ind.get("fiscal") is not None:
        score += 1 if ind["fiscal"] > 0 else 0
        parts.append(f"财政支出{ind['fiscal']:+.1f}%")
    if ind.get("bond_momentum") is not None:
        score += 1 if ind["bond_momentum"] > 0 else 0
        parts.append(f"专项债动量{ind['bond_momentum']:+.1f}%")
    if ind.get("corp_loan_chg") is not None:
        score += 1 if ind["corp_loan_chg"] > 0 else 0
        parts.append(f"企业贷款{'增' if ind['corp_loan_chg'] > 0 else '减'}")
    return {
        "label": "政策执行确认", "value": score, "forecast": None, "prev": None,
        "date": datetime.now(BEIJING).strftime("%Y-%m"), "hist": [],
        "unit": "/3", "source": "财政+专项债+信贷三线合成",
        "desc": "；".join(parts),
    }


# ---------------------------------------------------------------------------
# 汇总入口：一次拉全所有扩展指标
# ---------------------------------------------------------------------------

def fetch_all() -> dict:
    """并行拉取全部扩展指标，返回 {key: 指标卡}。单源失败不影响其他。"""
    from concurrent.futures import ThreadPoolExecutor
    specs = {
        # NBS
        "pmi_full": pmi_full,
        "core_cpi": core_cpi,
        "industrial_profit": industrial_profit,
        "industrial_inventory": industrial_inventory,
        "industrial_revenue": industrial_revenue,
        "property_sales_area": property_sales_area,
        "property_funds": property_funds,
        "property_loans": property_loans,
        "fai_equipment": fai_equipment,
        "fiscal_expenditure": fiscal_expenditure,
        # 央行
        "social_financing_stock": social_financing_stock,
        "credit_by_sector": credit_by_sector,
        "bank_survey": bank_survey,
        # 财政部
        "fiscal_revenue_expenditure": fiscal_revenue_expenditure,
        # CPB
        "world_trade_volume": world_trade_volume,
        # 专项债
        "special_bond_issuance": special_bond_issuance,
        "financial_conditions": financial_conditions,
        # 近似指标
        "copper_oil_ratio": copper_oil_ratio,
        "resale_house_breadth": resale_house_breadth,
        "services_production": services_production,
        "caixin_manufacturing_pmi": caixin_manufacturing_pmi,
        "caixin_services_pmi": caixin_services_pmi,
        "usdcnh": usdcnh,
        "nonbank_deposit": nonbank_deposit,
        "policy_execution": policy_execution,
    }
    out: dict = {}
    with ThreadPoolExecutor(max_workers=5) as pool:
        futures = {k: pool.submit(fn) for k, fn in specs.items()}
        for k, fut in futures.items():
            try:
                val = fut.result()
            except Exception:
                val = None
            if val is None:
                continue
            if k in ("pmi_full", "financial_conditions") and isinstance(val, dict):
                out.update(val)  # 展开 PMI 分项
            else:
                out[k] = val
    return out


if __name__ == "__main__":
    import json
    data = fetch_all()
    print(json.dumps({k: {kk: (vv if not isinstance(vv, list) else f"{len(vv)}pts")
                          for kk, vv in v.items() if kk in ("label", "value", "date", "hist", "unit", "source")}
                      for k, v in data.items()}, ensure_ascii=False, indent=2))
